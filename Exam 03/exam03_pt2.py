from openpyxl import load_workbook
import matplotlib.pyplot as plt
from scipy import optimize
import numpy as np

def f(theta, x):
    return theta*x*(1-x)

def plot_3d(grid_errors, name):
    plt.clf()
    z, x, y = zip(*grid_errors)
    ax = plt.axes(projection='3d')
    ax.scatter(x, y, z, c=z, cmap='viridis', linewidth=0.5)
    ax.set_xlabel("$x_0$")
    ax.set_ylabel("$\\theta$")
    ax.set_zlabel("$\\varepsilon$")
    plt.title(name)
    plt.savefig("{0}.png".format(name))

def norm_p(e, p):
    ans = None
    if p == np.inf: ans = max(e)
    else:
        sum = 0
        for x in e: sum += abs(x)**p
        ans = (sum **(1/p))
    return ans

def euler_explicit_method(t_vector, y0, h, theta):
    n = len(t_vector)
    e_vector = np.zeros(n)
    e_vector[0] = y0
    for i in range(n-1):
        e_vector[i + 1] = e_vector[i] + h*f(theta, e_vector[i])
    return e_vector

def euler_implicit_method(t_vector, y0, h, theta):
    n = len(t_vector)
    e_vector = np.zeros(n)
    e_vector[0] = y0
    for i in range(n-1):
        f_ = lambda x: x - e_vector[i] - (h * f(theta, x))
        root = optimize.fsolve((f_),(e_vector[i]))
        e_vector[i + 1] = root
    return e_vector

def trapezoid_method(t_vector, y0, h, theta):
    n = len(t_vector)
    e_vector = np.zeros(n)
    e_vector[0] = y0
    for i in range(n-1):
        f_ = lambda x: x - e_vector[i] - ((h/2)*(f(theta, e_vector[i]) + f(theta, x)))
        root = optimize.fsolve((f_),(e_vector[i]))
        e_vector[i + 1] = root
    return e_vector

n = 121
time = [x+1 for x in range(n)]
data = []
workbook = load_workbook(filename="mediciones.xlsx")
sheet = workbook.active
for column in sheet.iter_cols(min_row=2, max_row=2, min_col=2, max_col=122, values_only=True):
    data.append(column[0])

#Grid Search
N = 50 #Grid Size
x0_min, x0_max = 0, 1
x0_range = np.linspace(x0_min, x0_max, num=N)
theta_min, theta_max = 0, 1
theta_range = np.linspace(theta_min, theta_max, num=N)

grid_errors_ee = []
grid_errors_ei = []
grid_errors_t = []
for i in range (N):
    for j in range(N):
        x0 = x0_range[i]
        theta = theta_range[j]
        #Euler Explicit
        ans1 = euler_explicit_method(time, x0, 1, theta)
        e1 = np.subtract(ans1,data)
        grid_errors_ee.append((norm_p(e1, 2), x0, theta))
        #Euler Implicit
        ans2 = euler_implicit_method(time, x0, 1, theta)
        e2 = np.subtract(ans2,data)
        grid_errors_ei.append((norm_p(e2, 2), x0, theta))
        #Trapezoid
        ans3 = trapezoid_method(time, x0, 1, theta)
        e3 = np.subtract(ans3,data)
        grid_errors_t.append((norm_p(e3, 2), x0, theta))


best_parameters_ee = min(grid_errors_ee, key = lambda l: l[0])
print("El error mas pequeno encontrado con el metodo de euler explicito fue {0}, el cual se encontro con x0={1} y theta={2}".format(best_parameters_ee[0], best_parameters_ee[1], best_parameters_ee[2]))
plot_3d(grid_errors_ee, "euler explicit")

best_parameters_ei = min(grid_errors_ei, key = lambda l: l[0])
print("El error mas pequeno encontrado con el metodo de euler implicito fue {0}, el cual se encontro con x0={1} y theta={2}".format(best_parameters_ei[0], best_parameters_ei[1], best_parameters_ei[2]))
plot_3d(grid_errors_ee, "euler implicit")

best_parameters_t = min(grid_errors_t, key = lambda l: l[0])
print("El error mas pequeno encontrado con el metodo del trapecio fue {0}, el cual se encontro con x0={1} y theta={2}".format(best_parameters_t[0], best_parameters_t[1], best_parameters_t[2]))
plot_3d(grid_errors_ee, "trapezoid")
